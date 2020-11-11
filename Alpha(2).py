from openpyxl import load_workbook
from docx import Document
import re
import os

# 文件路径，我自己定义了一个，到时候要用用户的
file_path = 'C:\\Users\\admin\\Desktop\\学生信息表.xlsx'

# 简单的模板
data = "我的名字是{1},学号是{5},性别{3},年龄{2},在{4}学院学习"

# 选定用来命名输出文件的列
key = '学号'

# 获取文件名
# file_name = re.findall(r'[^\\/:*?"<>|\r\n]+$', file_path)
# file_name = re.findall(r'(.+?)\.xlsx', file_name[0])

try:
    # 读取excel xlsx文件
    wb = load_workbook(file_path)

    # 获取所有sheet页名字
    xl_sheet_names = wb.sheetnames

    # 定位到相应sheet页,[0]为sheet页索引
    ws = wb[xl_sheet_names[0]]

    # 获取行列数
    excel_row = ws.max_row

    # 找到模板中需要插入内容的位置
    blanks = re.findall(r'{\d+}', data)

    # 找到key对应的列
    k = 0
    for row in list(ws.rows)[0]:
        k += 1
        if key == row.value:
            break

    # i表示第i行数据
    i = 2

    # 写入word文件
    while i <= excel_row:
        tem = data
        for blank in blanks:
            index = int(blank[1:-1])  # 找到要填入的内容对应excel表格的单元格
            value = str(ws.cell(row=i, column=index).value)  # 找到要替换的内容
            tem = tem.replace(blank, value)  # 替换
        # 文件名
        name = str(ws.cell(row=i, column=k).value)
        # 创建word文档
        document = Document()
        # 向文档中写入内容
        document.add_paragraph(tem)
        # 输出路径的前面一部分到时候用用户输入的
        docx_path = 'C:\\Users\\admin\\Desktop\\' + name + '.docx'
        document.save(docx_path)
        i += 1

except Exception as err:
    print(err)